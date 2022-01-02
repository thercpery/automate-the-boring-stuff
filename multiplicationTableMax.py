#! python3

import logging, openpyxl
from openpyxl.styles import Font
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")

wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row = 1, column = 1).value = "MULTIPLICATION TABLE"
sheet.cell(row = 1, column = 1).font = Font(name = "Verdana", bold = True)
sheet.column_dimensions["A"].width = 35
for i in range(1, 18276):
    # Fill rows
    sheet.cell(row = i + 1, column = 1).value = i
    sheet.cell(row = i + 1, column = 1).font = Font(name = "Verdana", bold = True)
    # Fill columns
    sheet.cell(row = 1, column = i + 1).value = i
    sheet.cell(row = 1, column = i + 1).font = Font(name = "Verdana", bold = True)
    logging.debug(f"i = {i}")
    for j in range(1, 18276):
        logging.debug(f"j = {j}")
        product = i * j
        logging.debug(f"product = {i} * {j}")
        logging.debug(f"product = {product}")
        sheet.cell(row = i + 1, column = j + 1).value = product

# Finishing touches
sheet.freeze_panes = "C2"
print("Saving file...")
wb.save("multiplicationTableMax.xlsx")